import os
import glob
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

def get_pdf_pages(file_path):
    try:
        reader = PdfReader(file_path)
        return len(reader.pages)
    except Exception:
        return 0

def get_docx_pages(file_path):
    try:
        # Lưu ý: Số trang của docx được lấy từ metadata (thuộc tính app.xml)
        # Nó có thể không chính xác tuyệt đối nếu file chưa được lưu cập nhật lần cuối
        doc = Document(file_path)
        # Truy cập vào app properties để lấy số trang
        return doc.core_properties.page_count if doc.core_properties.page_count else 1
    except Exception:
        return 0

def get_xlsx_pages(file_path):
    try:
        # Excel không có khái niệm "trang" cố định cho đến khi in.
        # Ở đây ta sẽ tính 1 Sheet = 1 Trang.
        wb = load_workbook(file_path, read_only=True, data_only=True)
        return len(wb.sheetnames)
    except Exception:
        return 0

def calculate_average_pages(folder_path):
    total_pages = 0
    total_files = 0
    
    # Duyệt qua tất cả các file trong thư mục
    # Hỗ trợ cả file trong thư mục con (recursive) nếu muốn
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            ext = file.lower().split('.')[-1]
            
            pages = 0
            is_valid_file = False
            
            if ext == 'pdf':
                pages = get_pdf_pages(file_path)
                is_valid_file = True
            elif ext == 'docx':
                pages = get_docx_pages(file_path)
                is_valid_file = True
            elif ext == 'xlsx':
                pages = get_xlsx_pages(file_path)
                is_valid_file = True
            
            if is_valid_file and pages > 0:
                total_pages += pages
                total_files += 1
                print(f"Đã đọc: {file} - {pages} trang")

    if total_files == 0:
        return 0, 0

    avg_pages = total_pages / total_files
    return avg_pages, total_files

# --- Chạy thử ---
folder_path = input("Nhập đường dẫn thư mục: ")
if os.path.exists(folder_path):
    avg, count = calculate_average_pages(folder_path)
    print("-" * 30)
    print(f"Tổng số file xử lý được: {count}")
    print(f"Số trang trung bình: {avg:.2f}")
else:
    print("Đường dẫn thư mục không tồn tại.")